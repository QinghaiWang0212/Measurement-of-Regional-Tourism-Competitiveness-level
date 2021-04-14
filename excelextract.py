import math
import numpy as np
import xlrd
import openpyxl

# 从excel表中获取所有sheet的名字（使用xlrd方法）
def getSheet(excelfile):
    ExcelFile=xlrd.open_workbook(excelfile)
    sn = ExcelFile.sheet_names()
    sheet=[]
    for i in range(len(sn)):
        sheet.append(ExcelFile.sheet_by_name(sn[i]))
    return sheet
# 从excel表中获取所有sheet对象（使用xlrd方法）
# def get_sheet_by_name(excelfile,sheetname):
#     excelfile=xlrd.open_workbook(excelfile)
#     return  excelfile.sheet_by_name(sheetname)

def get_sheet_by_name(excelfile_formal_parameter,sheetname):
    excelfile=openpyxl.load_workbook(excelfile_formal_parameter)
    return  excelfile[sheetname]

def get_sheet_by_sequence(excelfile_formal_parameter,sequence):
    excelfile=openpyxl.load_workbook(excelfile_formal_parameter)
    return  excelfile[excelfile.sheetnames[sequence]]

# 从指定sheet中读取3,4,5,6,7,9,10,11,12,14,15,16,17行数据，从每行中读取3-14列的数据到矩阵data中。
# 其中sheet对象应为xlrd.open_workbook返回的sheet对象
def get_list_data_from_excel(sheet,result_rows=0,result_cols=0,start_row=1,start_col=0,selected_rows=[],selected_cols=[]):
    if not selected_rows :
        selected_rows = [i+start_row for i in range(result_rows)]

    if not selected_cols :
        selected_cols = [i+start_col for i in range(result_cols)]


    list_data = []
    temp_data = []
    r=0
    for row in selected_rows:
        # temp_data.append(sheet.row_values(row))  #使用xlrd方法读excel
        list_data.append([])
        for col in selected_cols:
            list_data[r].append(sheet.cell(row=row, column=col).value)
            # list_data[r].append(temp_data[r][col])  # 使用xlrd方法读excel
        r += 1
    return list_data

def get_array_data_from_excel(sheet,result_rows,result_cols,start_row=2,start_col=2):
    array_data = np.zeros((result_rows,result_cols))
    for i in range(array_data.shape[0]):
        for j in range(array_data.shape[1]):
            array_data[i][j] = sheet.cell(row=(i+start_row),column=(j+start_col)).value
    return array_data






# def getbiaozhunfromresult(sheet):
#     Data=np.zeros((13,11))
#     for i in range(1,14):
#         xrow = sheet.row_values(i)
#         xrow = xrow[1:12]
#         for j in range(len(xrow)):
#             Data[i - 1][j] = xrow[j]
#     return Data

# 从输出到result_xx的excel的权重sheet当中读数据到矩阵中
# def getweightfromresult(sheet):
#     Data =np.zeros((13,13))
#     for i in range(1,14):
#         xrow = sheet.row_values(i)
#         xrow = xrow[1:14]
#         for j in range(len(xrow)):
#             Data[i-1][j]  = xrow[j]
#     return  Data

# def biaozhunhua(Data):
#     biao = np.zeros((13, 11))
#     for i in range(Data.shape[0]):
#         row = Data[i, :]
#         m = np.max(row)
#         n = np.min(row)
#         for j in range(len(row)):
#             if (m - n) == 0:
#                 biao[i][j] = 0
#             else:
#                 biao[i][j] = (row[j]-n) / (m - n)
#     return biao



# 不带区间扩展的标准化（一般标准化）


def standardize(row):
    if (row[len(row)-1] == "p"):
        # print("p")
        return standardize_positive(np.array(row[:-1]).astype(np.float))
    elif (row[len(row)-1] == "n"):
        # print("n")
        return standardize_negative(np.array(row[:-1]).astype(np.float))




def standardize_negative(row):
    # print(row)
    standardized_data = np.zeros(len(row))
    max = np.max(row)
    min = np.min(row)
    for i in range(len(row)):
        if (max - min) == 0:
            standardized_data[i] = 0
        else:
            standardized_data[i] = (max - row[i]) /(max - min)
    # 使标准化数据中没有0：
    # for i in range(len(standardized_data)):
    #     if standardized_data[i]==0:
    #         standardized_data = standardized_data + 0.0001
    #         break



    return standardized_data


def standardize_positive(row):
    # print(row)
    standardized_data = np.zeros(len(row))
    max = np.max(row)
    min = np.min(row)
    for i in range(len(row)):
        if (max - min) == 0:
            standardized_data[i] = 0
        else:
            standardized_data[i] = (row[i] - min) /(max - min)
    # 使标准化数据中没有0：
    # for i in range(len(standardized_data)):
    #     if standardized_data[i]==0:
    #         standardized_data = standardized_data + 0.0001
    #         break

    return standardized_data






# 基于区间扩展的标准化

def standardize_extend_interval(row,k=0.05):
    if (row[len(row)-1] == "p"):
        return standardize_positive_extend_interval(np.array(row[:-1]).astype(np.float),k)
    else:
        return standardize_negative_extend_interval(np.array(row[:-1]).astype(np.float),k)



def standardize_negative_extend_interval(row,k = 0.05):
    standardized_data = np.zeros(len(row))
    max = np.max(row)
    min = np.min(row)
    max_extend = max + max  * k
    min_extend = min -  min * k
    for i in range(len(row)):
        if (max_extend - min_extend) == 0:
            standardized_data[i] = 0
        else:
            standardized_data[i] = (max_extend - row[i]) /(max_extend - min_extend)
    for i in range(len(standardized_data)):
        if standardized_data[i]==0:
            standardized_data = standardized_data + 0.0001
            break
    return standardized_data


def standardize_positive_extend_interval(row,k=0.05):
    standardized_data = np.zeros(len(row))
    max = np.max(row)
    min = np.min(row)
    max_extend = max + max * k
    min_extend = min -  min * k
    for i in range(len(row)):
        if (max_extend - min_extend) == 0:
            standardized_data[i] = 0
        else:
            standardized_data[i] = (row[i] - min_extend) /(max_extend - min_extend)
    for i in range(len(standardized_data)):
        if standardized_data[i]==0:
            standardized_data = standardized_data + 0.0001
            break
    return standardized_data




def get_3D_array(storey):

    #
    city_sheet_collector = []
    # 三维矩阵，其中每个二维矩阵为一个城市的面板数据


    for i in range(len(result_city)):

        city_data = openpyxl.load_workbook(result_city[i])
        city_sheet_collector.append(city_data[sheet_name])
    for i in range(len(result_city)):
        for j in range(rows):
            for k in range(cols):
                D_array[i][j][k]=city_sheet_collector[i][j+start_row][k+start_col].value


    return  D_array




def get_deep_array(multi_city_array,row,col):
    deep_array = np.zeros(multi_city_array.shape[0])
    for i in range(len(deep_array)):
        deep_array[i] = multi_city_array[i][row][col]
    return deep_array



def standardize_multi_city_array(multi_city_array,k = 0.05):

    standardized_multi_city_array = np.zeros(multi_city_array.shape)
    for i in range(multi_city_array.shape[1]):
        for j in range(multi_city_array.shape[2]):
            deep_array = get_deep_array(multi_city_array,i,j)
            standardized_extend_deep_array = standardrize_positive_extend_interval(deep_array,k)
            for m in range(multi_city_array.shape[0]):
                standardized_multi_city_array[m][i][j] = standardized_extend_deep_array[m]
    return standardized_multi_city_array

# def write_two_dimensional_list_to_excel(two_dimensional_list,excel_name):
#     for i in range(len(two_dimensional_list)):
#         for j in range(len(two_dimensional_list[i])):



def entropy_weight_method(index_matrix):
    p = np.zeros((index_matrix.shape[0],index_matrix.shape[1]))

    e = np.zeros(index_matrix.shape[0])
    d = np.zeros(index_matrix.shape[0])
    q = np.zeros(index_matrix.shape[0])
    # 将0值标准化数据整行+0.0001平移
    for i in range(index_matrix.shape[0]):
        for j in range(index_matrix.shape[1]):
            if index_matrix[i][j] == 0:
                index_matrix[i,:] = index_matrix[i,:]+0.0001
                break



    for i in range(p.shape[0]):
        for j in range(p.shape[1]):
            # 按行求和(求每个指标下所有对象之和)，求每个对象在其指标中占的比重
            if index_matrix.sum(axis=1)[i]==0:
                p[i][j]=0
            else:
                p[i][j] = index_matrix[i][j]/index_matrix.sum(axis=1)[i]

    # 求每个指标的熵值


    for i in range(p.shape[0]):
        temp = 0
        for j in range(p.shape[1]):

            if p[i][j] == 0:
                ml = 0
            else:
                ml = math.log(p[i][j])
            temp = p[i][j] * ml +temp
        e[i] = (-1/math.log(index_matrix.shape[1])) * temp

    # 求信息熵冗余度
    for i in range(p.shape[0]):
        d[i] = 1 - e[i]

    # 求各项指标的权值
    for i in range(p.shape[0]):
        q[i] = d[i] / d.sum()
    # # 将中间值输出到  熵权法中间数据.xlsx
    # write_index_matrix = pd.DataFrame(index_matrix)
    # write_p = pd.DataFrame(p)
    # write_e = pd.DataFrame(e)
    # write_d = pd.DataFrame(d)
    # write_q = pd.DataFrame(q)
    # writer = pd.ExcelWriter("熵权法中间数据.xlsx")
    # write_p.to_excel(writer,startrow=1)
    # write_e.to_excel(writer,startrow=20)
    # write_d.to_excel(writer,startrow=40)
    # write_q.to_excel(writer,startrow=60)
    # write_index_matrix.to_excel(writer,startcol=20)
    # writer.save()
    return  q


def Sigmoid(x):
    s = 1 / (1 + np.exp(-x))
    return s







# 将结果矩阵输出到excel
"""
workbook = xlwt.Workbook(encoding = 'utf-8')
for i in range(8):
    w=getData(sheet[i])
    worksheet = workbook.add_sheet('cityx'+str(i))
    for j in range(w.shape[0]):
        for k in range(w.shape[1]):
            worksheet.write(j, k, w[j][k])
workbook.save('Excel_test.xlsx')
"""
